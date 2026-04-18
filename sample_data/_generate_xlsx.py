"""
One-off script to generate the sample .xlsx files for sample_data/.
Run from project root: uv run python sample_data/_generate_xlsx.py
Requires: openpyxl (already a transitive dependency).
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl import Workbook

BASE = os.path.dirname(os.path.abspath(__file__))


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def make_header_row(ws, headers, row=1):
    """Write bold, coloured header row."""
    fill = PatternFill("solid", fgColor="2B5290")
    font = Font(bold=True, color="FFFFFF")
    for col_idx, val in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")


# ---------------------------------------------------------------------------
# 1. Groww MF Excel  (skiprows=11 in ingest_mf.py)
#    Groww exports have 11 rows of metadata before the real header.
# ---------------------------------------------------------------------------

def create_groww_mf_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "MF Orders"

    # 11-row preamble that Groww places before data headers
    ws["A1"] = "Groww Mutual Fund Orders Report"
    ws["A2"] = "Account holder: Rahul Sample"
    ws["A3"] = "Downloaded: 01-Apr-2025"
    for r in range(4, 12):
        ws.cell(row=r, column=1, value="")   # empty filler rows

    # Row 12 = actual header (row index 11 when 0-based, skiprows=11 skips rows 1-11)
    headers = [
        "Scheme Name", "AMFI", "Date", "Transaction Type",
        "Amount", "Units", "NAV", "Status",
    ]
    make_header_row(ws, headers, row=12)

    data = [
        ["Parag Parikh Flexi Cap Fund - Direct Plan - Growth",     "122639", "15-Jan-2023", "Purchase",  10000, 12.345, 51.23, "Executed"],
        ["Parag Parikh Flexi Cap Fund - Direct Plan - Growth",     "122639", "15-Feb-2023", "Purchase",  10000, 11.987, 52.11, "Executed"],
        ["Parag Parikh Flexi Cap Fund - Direct Plan - Growth",     "122639", "15-Mar-2023", "Purchase",  10000, 10.523, 54.67, "Executed"],
        ["Axis Small Cap Fund - Direct Plan - Growth",             "125354", "10-Jan-2023", "Purchase",  15000, 18.234, 82.34, "Executed"],
        ["Axis Small Cap Fund - Direct Plan - Growth",             "125354", "10-Apr-2023", "Purchase",  15000, 17.812, 84.21, "Executed"],
        ["Mirae Asset Focused Fund Direct Plan Growth",            "147206", "20-Feb-2023", "Purchase",  20000, 32.100, 62.30, "Executed"],
        ["Mirae Asset Focused Fund Direct Plan Growth",            "147206", "20-May-2023", "Purchase",  20000, 31.450, 63.58, "Executed"],
        ["Mirae Asset Focused Fund Direct Plan Growth",            "147206", "05-Aug-2023", "Redemption",5000,  7.200, 69.44, "Executed"],
        ["quant Small Cap Fund - Growth Option - Direct Plan",     "120828", "01-Mar-2023", "Purchase",  8000,   6.512,122.85, "Executed"],
        ["quant Small Cap Fund - Growth Option - Direct Plan",     "120828", "01-Jun-2023", "Purchase",  8000,   5.876,136.15, "Executed"],
        ["Cancelled Order Example Fund",                           "",       "12-Dec-2023", "Purchase",  5000,    None,  None, "Rejected"],
    ]
    for r_idx, row in enumerate(data, start=13):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    path = os.path.join(BASE, "mf", "Rahul_MF.xlsx")
    wb.save(path)
    print(f"Created: {path}")


# ---------------------------------------------------------------------------
# 2. Groww Stock Excel  (skiprows=5 in ingest_stocks.py)
# ---------------------------------------------------------------------------

def create_groww_stock_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Orders"

    ws["A1"] = "Groww Equity Orders Report"
    ws["A2"] = "Account holder: Rahul Sample"
    ws["A3"] = "Downloaded: 01-Apr-2025"
    for r in range(4, 6):
        ws.cell(row=r, column=1, value="")

    headers = [
        "Stock name", "Symbol", "Type", "Quantity", "Value",
        "Execution date and time", "Order status",
    ]
    make_header_row(ws, headers, row=6)

    data = [
        ["Infosys Limited",            "INFY",      "BUY",  10, 148320.00, "15-Jan-2023 10:12:34", "EXECUTED"],
        ["Infosys Limited",            "INFY",      "BUY",   5,  76145.00, "10-Jun-2023 11:45:10", "EXECUTED"],
        ["HDFC Bank Limited",          "HDFCBANK",  "BUY",  20, 295800.00, "20-Feb-2023 14:30:22", "EXECUTED"],
        ["HDFC Bank Limited",          "HDFCBANK",  "SELL",  5,  75200.00, "05-Sep-2023 09:55:00", "EXECUTED"],
        ["Tata Consultancy Services",  "TCS",       "BUY",   8, 289600.00, "03-Mar-2023 13:22:11", "EXECUTED"],
        ["Zydus Wellness Limited",     "ZYDUSWELL", "BUY",  15, 247500.00, "22-Apr-2023 10:01:05", "EXECUTED"],
        ["Narayana Hrudayalaya Ltd.",  "NH",        "BUY",  50, 556000.00, "14-Jul-2023 11:11:11", "EXECUTED"],
        ["Cancelled Stock",            "FAKE",      "BUY",   1,   5000.00, "01-Jan-2024 09:00:00", "REJECTED"],
    ]
    for r_idx, row in enumerate(data, start=7):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    path = os.path.join(BASE, "stock", "Rahul_stocks.xlsx")
    wb.save(path)
    print(f"Created: {path}")


# ---------------------------------------------------------------------------
# 3. FD details Excel
# ---------------------------------------------------------------------------

def create_fd_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "FD Details"

    headers = ["Owner", "FD Start Date", "Maturity Date", "Invested Amount", "Interest Rate", "Bank"]
    make_header_row(ws, headers, row=1)

    data = [
        ["Rahul", "01-04-2023", "01-04-2025", 200000, 7.10, "State Bank of India"],
        ["Rahul", "15-09-2023", "15-09-2025", 100000, 7.25, "HDFC Bank"],
        ["Priya", "01-01-2024", "01-01-2026", 150000, 7.50, "ICICI Bank"],
    ]
    for r_idx, row in enumerate(data, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    path = os.path.join(BASE, "FD", "FD_details.xlsx")
    wb.save(path)
    print(f"Created: {path}")


# ---------------------------------------------------------------------------
# Run all generators
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    create_groww_mf_xlsx()
    create_groww_stock_xlsx()
    create_fd_xlsx()
    print("\nAll sample Excel files generated successfully.")
